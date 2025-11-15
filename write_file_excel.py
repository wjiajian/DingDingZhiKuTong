import os
import openpyxl
import docx
from openpyxl.utils import get_column_letter
from xbot import print

# --- 模块化的内容读取区域 ---
# 未来若要添加对新文件类型（例如 .csv）的支持:
# 1. 编写一个新的函数 `read_csv_content(file_path)`。
# 2. 在 FILE_READERS 字典中增加一行映射：`'.csv': read_csv_content`。

def read_txt_content(file_path: str) -> str:
    """从 .txt 文件中读取内容。"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"读取 TXT 时出错: {e}"

def read_docx_content(file_path: str) -> str:
    """从 .docx 文件中读取内容。"""
    try:
        doc = docx.Document(file_path)
        full_text = [para.text for para in doc.paragraphs]
        return '\n'.join(full_text)
    except Exception as e:
        return f"读取 DOCX 时出错: {e}"

def read_xlsx_content(file_path: str) -> str:
    """
    从 .xlsx 文件中的所有工作表读取可见的文本内容。
    """
    try:
        # 以只读模式加载工作簿，这样性能更好，且不会意外修改文件
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        all_sheets_text = []

        # 遍历工作簿中的每一个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []
            
            # 添加工作表标题，以便区分不同工作表的内容
            sheet_text.append(f"--- 工作表: {sheet.title} ---")

            # 遍历工作表中的每一行
            for row in sheet.iter_rows():
                # 获取行中每个单元格的值，并转换为字符串，忽略空单元格
                # str(cell.value) 可以安全地处理数字、日期等不同类型
                row_values = [str(cell.value) for cell in row if cell.value is not None]
                
                # 如果行中有内容，则将它们用制表符连接起来
                if row_values:
                    sheet_text.append("\t".join(row_values))
            
            # 将当前工作表的所有文本行用换行符连接起来
            all_sheets_text.append("\n".join(sheet_text))
        
        # 将所有工作表的内容用两个换行符隔开，使其更清晰
        return "\n\n".join(all_sheets_text)

    except FileNotFoundError:
        return f"错误：Excel 文件未找到 '{file_path}'"
    except Exception as e:
        return f"读取 XLSX 时出错: {e}"

# 这是分发字典，它将文件扩展名映射到正确的读取函数。
FILE_READERS = {
    '.txt': read_txt_content,
    '.docx': read_docx_content,
    '.xlsx': read_xlsx_content,
    # 在这里添加新的读取函数，例如: '.pdf': read_pdf_content
}

def get_content_from_file(file_path: str) -> str:
    """
    从文件中获取内容的通用函数。
    它使用 FILE_READERS 字典来查找并调用正确的读取器。
    """
    if not os.path.exists(file_path):
        return "错误：链接的文件不存在"
    
    # 获取文件的扩展名
    _, extension = os.path.splitext(file_path)
    
    # 在我们的字典中查找对应的读取函数
    reader_func = FILE_READERS.get(extension.lower())
    
    if reader_func:
        # 如果找到了读取函数，就调用它
        return reader_func(file_path)
    else:
        # 否则，返回不支持的类型错误
        return f"错误：不支持的文件类型 ({extension})"

# --- 主 Excel 处理逻辑 ---

def process_excel_in_place(excel_path: str):
    """
    自动查找链接列，在其后插入一个新列，
    用链接文档的内容填充它，并直接在原文件上保存更改。
    """
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        print(f"成功加载文件: '{excel_path}'")
    except Exception as e:
        print(f"加载 Excel 文件时出错: {e}")
        return

    # 获取Excel文件所在的绝对目录
    excel_base_dir = os.path.dirname(os.path.abspath(excel_path))
    print(f"将基于此目录解析相对路径: '{excel_base_dir}'")

    all_links = [{'cell': cell, 'target': cell.hyperlink.target}
                 for row in sheet.iter_rows() for cell in row if cell.hyperlink]

    if not all_links:
        print("在此文件中未找到任何超链接。未做任何更改。")
        return
    
    print(f"找到了 {len(all_links)} 个超链接。")

    first_link_col_idx = all_links[0]['cell'].column
    content_col_idx = first_link_col_idx + 1
    
    print(f"检测到链接列为 {get_column_letter(first_link_col_idx)} 列。 "
          f"将在 {get_column_letter(content_col_idx)} 列插入新内容。")
          
    sheet.insert_cols(content_col_idx)
    
    header_cell = sheet.cell(row=1, column=content_col_idx)
    header_cell.value = "链接文档内容"
    header_cell.font = openpyxl.styles.Font(bold=True)

    for link_info in all_links:
        link_cell = link_info['cell']
        # 这是从Excel中读取的原始路径，可能是相对的
        relative_or_absolute_path = link_info['target']
        
        # 解析路径，将相对路径转换为绝对路径
        if os.path.isabs(relative_or_absolute_path):
            # 如果路径已经是绝对路径 (例如 "C:\...")，则直接使用
            full_path = relative_or_absolute_path
        else:
            # 如果是相对路径，则与Excel文件所在目录进行拼接
            full_path = os.path.join(excel_base_dir, relative_or_absolute_path)
            
        print(f"  - 正在处理 {link_cell.coordinate}: '{relative_or_absolute_path}' -> 解析为 '{full_path}'")
        
        # 使用解析后的完整路径来获取内容
        content = get_content_from_file(full_path)
        
        content_cell = sheet.cell(row=link_cell.row, column=content_col_idx)
        content_cell.value = content

    try:
        print(f"\n正在将更改保存到原始文件: '{excel_path}'...")
        workbook.save(excel_path)
        print("处理完成！原始文件已更新。")
    except PermissionError:
        print(f"\n错误：无法保存文件。请确保 '{excel_path}' 没有被其他程序（如Excel）打开。")
    except Exception as e:
        print(f"\n保存文件时发生未知错误: {e}")

# --- 脚本主入口 ---
if __name__ == "__main__":
    # --- 警告 ---
    # 此脚本将直接修改您的原始文件。
    # 强烈建议在运行前对您的 Excel 文件进行备份。
    
    # --- 请在这里提供您的 Excel 文件的完整路径 ---
    excel_file_path = "file_path"
    
    process_excel_in_place(excel_file_path)
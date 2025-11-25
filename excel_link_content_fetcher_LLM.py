# -*- coding: utf-8 -*-
"""
Excel链接内容提取器 - LLM增强版

基于Unstructured和多模态大模型的增强型Excel超链接文档内容提取工具，支持30+种文档格式，
通过LLM高精度的图片内容识别和解析。
具备错误恢复、性能优化等企业级特性。
"""

import os
import shutil
import time
import logging
import mimetypes
import tempfile
import base64
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

try:
    from unstructured.partition.auto import partition

    UNSTRUCTURED_AVAILABLE = True
except ImportError:
    UNSTRUCTURED_AVAILABLE = False
    partition = None

# 1. 屏蔽 pdfminer 的字体警告
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# 2. 屏蔽 huggingface 的下载警告和超时设置
logging.getLogger("huggingface_hub").setLevel(logging.ERROR)

# ==================== LLM多模态配置和数据结构 ====================


@dataclass
class MultimodalConfig:
    """多模态LLM处理配置"""

    # LLM模型配置
    vision_model_provider: str = "openai"  # openai, qwen
    vision_model_name: str = (
        "gpt-4-vision-preview"  # 或 "qwen-vl-flash" (通过OpenAI兼容方式调用)
    )
    api_key: Optional[str] = None
    api_base: Optional[str] = None

    # 图片处理配置
    max_image_size: int = 10 * 1024 * 1024  # 10MB
    supported_image_formats: List[str] = field(
        default_factory=lambda: [".png", ".jpg", ".jpeg", ".gif", ".bmp"]
    )
    extract_embedded_images: bool = True

    vision_prompt: str = "请详细识别和描述这张图片中的所有文本内容，包括标题、正文、表格、图表等，保持原有的格式和结构。如果是表格，请用markdown表格格式输出。"

    # 批处理配置
    batch_size: int = 5
    max_retries: int = 3
    retry_delay: float = 1.0

    # 成本控制
    max_tokens_per_image: int = 2000
    enable_caching: bool = True


@dataclass
class ImageContentInfo:
    """图片内容信息"""

    image_id: int
    source_path: str
    processed_content: str
    alt_text: str = ""
    extraction_type: str = "embedded"  # embedded, external, screenshot
    metadata: Dict[str, Any] = field(default_factory=dict)


# ==================== 原有配置和数据结构 ====================


@dataclass
class DocumentContent:
    """文档内容结构"""

    sections: List[Dict[str, Any]]
    metadata: Dict[str, Any]
    full_text: str

    def __post_init__(self):
        """后处理验证"""
        if not isinstance(self.sections, list):
            self.sections = []
        if not isinstance(self.metadata, dict):
            self.metadata = {}
        if not isinstance(self.full_text, str):
            self.full_text = ""


@dataclass
class DocumentProcessingConfig:
    """文档处理配置"""

    partition_strategy: str = "fast"  # hi_res, fast
    model_name: Optional[str] = None
    languages: List[str] = field(default_factory=lambda: ["zh", "en"])

    # 网络配置
    offline_mode: bool = False  # 是否使用离线模式
    max_download_retries: int = 5  # 最大重试次数
    download_timeout: int = 60  # 下载超时时间（秒）

    # 输出配置
    output_format: str = "markdown"  # markdown, plain_text, json
    max_content_length: int = 10000
    include_metadata: bool = True

    # 安全配置
    # backup_enabled: bool = True
    # backup_location: str = "./backup/"
    continue_on_error: bool = True
    max_retries: int = 3

    # 性能配置
    enable_parallel_processing: bool = True
    max_workers: int = 4
    memory_limit_mb: int = 512

    # LLM多模态配置
    enable_multimodal: bool = False
    multimodal_config: Optional[MultimodalConfig] = None

    def __post_init__(self):
        """后处理初始化"""
        # 确保languages不为空
        if not self.languages:
            self.languages = ["zh", "en"]

        # 验证输出格式
        valid_formats = ["markdown", "plain_text", "json"]
        if self.output_format not in valid_formats:
            self.output_format = "markdown"

        # 验证分区策略
        valid_strategies = ["hi_res", "fast"]
        if self.partition_strategy not in valid_strategies:
            self.partition_strategy = "fast"

        # 验证多模态配置
        if self.enable_multimodal and not self.multimodal_config:
            self.multimodal_config = MultimodalConfig()


@dataclass
class HyperlinkInfo:
    """超链接信息"""

    cell: openpyxl.cell.Cell
    target: str
    display_text: Union[str, int, float]

    def __post_init__(self):
        """后处理验证"""
        if self.display_text is None:
            self.display_text = ""


# ==================== LLM多模态处理器 ====================


class MultimodalProcessor:
    """多模态大模型处理器，用于高精度图片内容解析"""

    def __init__(self, config: MultimodalConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        self._cache = {} if config.enable_caching else None

    def extract_and_process_images(self, file_path: str) -> List[ImageContentInfo]:
        """
        提取并处理文档中的图片

        Args:
            file_path: 文档文件路径

        Returns:
            List[ImageContentInfo]: 处理后的图片内容信息列表
        """
        try:
            # 1. 提取嵌入图片
            extracted_images = self._extract_embedded_images(file_path)

            # 2. 处理外部图片引用
            external_images = self._extract_external_images(file_path)

            # 3. 合并所有图片
            all_images = extracted_images + external_images

            # 4. 批量处理图片
            processed_images = []
            for img_info in all_images:
                try:
                    processed_content = self._process_single_image(
                        img_info["source_path"]
                    )
                    processed_images.append(
                        ImageContentInfo(
                            image_id=len(processed_images),
                            source_path=img_info["source_path"],
                            processed_content=processed_content,
                            alt_text=img_info.get("alt_text", ""),
                            extraction_type=img_info.get("type", "embedded"),
                        )
                    )
                except Exception as e:
                    self.logger.error(
                        f"处理图片失败 {img_info.get('source_path', 'unknown')}: {e}"
                    )
                    processed_images.append(
                        ImageContentInfo(
                            image_id=len(processed_images),
                            source_path=img_info.get("source_path", "unknown"),
                            processed_content=f"[图片处理失败: {str(e)}]",
                            alt_text=img_info.get("alt_text", ""),
                            extraction_type=img_info.get("type", "embedded"),
                        )
                    )

            return processed_images

        except Exception as e:
            self.logger.error(f"图片提取和处理失败: {e}")
            return []

    def _extract_embedded_images(self, file_path: str) -> List[Dict[str, Any]]:
        """提取文档中的嵌入图片"""
        images = []

        try:
            _, extension = os.path.splitext(file_path.lower())

            if extension == ".pdf":
                images = self._extract_pdf_images(file_path)
            elif extension in [".docx", ".pptx", ".xlsx"]:
                images = self._extract_office_images(file_path, extension)

        except Exception as e:
            self.logger.error(f"提取嵌入图片失败: {e}")

        return images

    def _extract_pdf_images(self, pdf_path: str) -> List[Dict[str, Any]]:
        """提取PDF中的图片"""
        try:
            # 尝试使用PyMuPDF提取图片
            import fitz  # PyMuPDF

            images = []
            doc = fitz.open(pdf_path)

            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                image_list = page.get_images(full=True)

                for img_index, img in enumerate(image_list):
                    try:
                        xref = img[0]
                        pix = fitz.Pixmap(doc, xref)

                        if pix.n - pix.alpha < 4:  # GRAY or RGB
                            temp_path = f"{pdf_path}_img_{page_num}_{img_index}.png"
                            pix.save(temp_path)

                            images.append(
                                {
                                    "source_path": temp_path,
                                    "alt_text": f"PDF图片 - 页面{page_num + 1}, 图片{img_index + 1}",
                                    "type": "pdf_embedded",
                                }
                            )
                    except Exception as e:
                        self.logger.warning(f"提取PDF图片失败: {e}")

            doc.close()
            return images

        except ImportError:
            self.logger.warning("未安装PyMuPDF，跳过PDF图片提取")
            return []
        except Exception as e:
            self.logger.error(f"PDF图片提取失败: {e}")
            return []

    def _extract_office_images(
        self, file_path: str, extension: str
    ) -> List[Dict[str, Any]]:
        """提取Office文档中的图片"""
        images = []

        try:
            if extension == ".docx":
                # 使用python-docx提取图片
                try:
                    from docx import Document
                    import zipfile

                    # 方法1: 通过zipfile直接提取图片（更可靠）
                    with zipfile.ZipFile(file_path, "r") as zip_ref:
                        for file_info in zip_ref.namelist():
                            if file_info.startswith("word/media/") and any(
                                file_info.lower().endswith(ext)
                                for ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]
                            ):
                                try:
                                    img_data = zip_ref.read(file_info)
                                    img_name = os.path.basename(file_info)
                                    temp_path = os.path.join(
                                        tempfile.gettempdir(),
                                        f"{os.path.basename(file_path)}_{img_name}",
                                    )
                                    with open(temp_path, "wb") as f:
                                        f.write(img_data)

                                    images.append(
                                        {
                                            "source_path": temp_path,
                                            "alt_text": f"Word图片 - {img_name}",
                                            "type": "docx_embedded",
                                        }
                                    )
                                except Exception as e:
                                    self.logger.warning(
                                        f"提取Word图片失败 {file_info}: {e}"
                                    )

                except ImportError:
                    self.logger.warning("未安装python-docx，跳过Word文档图片提取")

            elif extension == ".pptx":
                # 使用zipfile直接提取图片（更可靠）
                try:
                    import zipfile

                    with zipfile.ZipFile(file_path, "r") as zip_ref:
                        for file_info in zip_ref.namelist():
                            if file_info.startswith("ppt/media/") and any(
                                file_info.lower().endswith(ext)
                                for ext in [
                                    ".png",
                                    ".jpg",
                                    ".jpeg",
                                    ".gif",
                                    ".bmp",
                                    ".tiff",
                                ]
                            ):
                                try:
                                    img_data = zip_ref.read(file_info)
                                    img_name = os.path.basename(file_info)
                                    temp_path = os.path.join(
                                        tempfile.gettempdir(),
                                        f"{os.path.basename(file_path)}_{img_name}",
                                    )
                                    with open(temp_path, "wb") as f:
                                        f.write(img_data)

                                    images.append(
                                        {
                                            "source_path": temp_path,
                                            "alt_text": f"PPT图片 - {img_name}",
                                            "type": "pptx_embedded",
                                        }
                                    )
                                except Exception as e:
                                    self.logger.warning(
                                        f"提取PPT图片失败 {file_info}: {e}"
                                    )

                except Exception as e:
                    self.logger.error(f"PowerPoint文档图片提取失败: {e}")

            elif extension == ".xlsx":
                # 使用zipfile直接提取Excel图片
                try:
                    import zipfile

                    with zipfile.ZipFile(file_path, "r") as zip_ref:
                        for file_info in zip_ref.namelist():
                            if file_info.startswith("xl/media/") and any(
                                file_info.lower().endswith(ext)
                                for ext in [
                                    ".png",
                                    ".jpg",
                                    ".jpeg",
                                    ".gif",
                                    ".bmp",
                                    ".tiff",
                                ]
                            ):
                                try:
                                    img_data = zip_ref.read(file_info)
                                    img_name = os.path.basename(file_info)
                                    temp_path = os.path.join(
                                        tempfile.gettempdir(),
                                        f"{os.path.basename(file_path)}_{img_name}",
                                    )
                                    with open(temp_path, "wb") as f:
                                        f.write(img_data)

                                    images.append(
                                        {
                                            "source_path": temp_path,
                                            "alt_text": f"Excel图片 - {img_name}",
                                            "type": "xlsx_embedded",
                                        }
                                    )
                                except Exception as e:
                                    self.logger.warning(
                                        f"提取Excel图片失败 {file_info}: {e}"
                                    )

                except Exception as e:
                    self.logger.error(f"Excel文档图片提取失败: {e}")

        except Exception as e:
            self.logger.error(f"Office文档图片提取失败: {e}")

        return images

    def _extract_external_images(self, file_path: str) -> List[Dict[str, Any]]:
        """提取外部引用的图片（截图、附件等）"""
        # 这里可以扩展处理外部图片文件
        # 比如同一目录下的图片文件，或文档中引用的外部图片链接
        return []

    def _process_single_image(self, image_path: str) -> str:
        """使用多模态模型处理单张图片"""
        try:
            # 检查缓存
            if self._cache and image_path in self._cache:
                return self._cache[image_path]

            # 检查文件是否存在
            if not os.path.exists(image_path):
                return f"[图片文件不存在: {image_path}]"

            # 检查文件大小
            file_size = os.path.getsize(image_path)
            if file_size > self.config.max_image_size:
                return f"[图片文件过大: {file_size} bytes, 最大允许: {self.config.max_image_size} bytes]"

            # 根据模型提供商选择处理方式
            if self.config.vision_model_provider == "openai":
                result = self._process_with_openai_vision(image_path)
            elif self.config.vision_model_provider == "qwen":
                result = self._process_with_qwen_vision(image_path)
            else:
                result = f"[不支持的模型提供商: {self.config.vision_model_provider}]"

            # 缓存结果
            if self._cache:
                self._cache[image_path] = result

            return result

        except Exception as e:
            return f"[图片处理失败: {str(e)}]"

    def _process_with_openai_vision(self, image_path: str) -> str:
        """使用OpenAI兼容的大模型处理图片"""
        try:
            from openai import OpenAI

            # 创建客户端
            client_params = {"api_key": self.config.api_key}
            if self.config.api_base:
                client_params["base_url"] = self.config.api_base

            client = OpenAI(**client_params)

            # 读取图片并转换为base64
            with open(image_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode("utf-8")

            # 调用API
            response = client.chat.completions.create(
                model=self.config.vision_model_name,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self.config.vision_prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{encoded_image}",
                                    "detail": "high",  # 高精度模式
                                },
                            },
                        ],
                    }
                ],
                max_tokens=self.config.max_tokens_per_image,
            )

            return response.choices[0].message.content

        except ImportError:
            return "[错误: 请安装openai库: pip install openai]"
        except Exception as e:
            return f"[OpenAI处理失败: {str(e)}]"

    def _process_with_qwen_vision(self, image_path: str) -> str:
        """使用通义千问Qwen-VL通过OpenAI兼容方式处理图片"""
        try:
            from openai import OpenAI

            # 配置OpenAI兼容的客户端
            client_params = {"api_key": self.config.api_key}

            # 设置DashScope的OpenAI兼容API端点
            if not self.config.api_base:
                self.config.api_base = (
                    "https://dashscope.aliyuncs.com/compatible-mode/v1"
                )

            client_params["base_url"] = self.config.api_base

            client = OpenAI(**client_params)

            # 读取图片并转换为base64
            with open(image_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode("utf-8")

            # 调用OpenAI兼容的Qwen-VL API
            response = client.chat.completions.create(
                model=self.config.vision_model_name,  # 例如: "qwen-vl-plus"
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self.config.vision_prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{encoded_image}"
                                },
                            },
                        ],
                    }
                ],
                max_tokens=self.config.max_tokens_per_image,
                temperature=0.1,
            )

            return response.choices[0].message.content

        except ImportError:
            return "[错误: 请安装openai库: pip install openai]"
        except Exception as e:
            return f"[Qwen处理失败: {str(e)}]"

    def cleanup_temporary_files(self, image_contents: List[ImageContentInfo]):
        """清理临时文件"""
        for img_info in image_contents:
            try:
                if img_info.source_path.startswith(tempfile.gettempdir()):
                    if os.path.exists(img_info.source_path):
                        os.remove(img_info.source_path)
            except Exception as e:
                self.logger.warning(f"清理临时文件失败 {img_info.source_path}: {e}")


# ==================== 原有核心处理模块 ====================


class UnifiedDocumentProcessor:
    """基于Unstructured的统一文档处理器，支持LLM多模态增强"""

    def __init__(self, config: DocumentProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        self.partition_options = self._build_partition_options()

        # 初始化多模态处理器
        if self.config.enable_multimodal and self.config.multimodal_config:
            self.multimodal_processor = MultimodalProcessor(
                self.config.multimodal_config
            )
        else:
            self.multimodal_processor = None

        # 预下载模型（如果需要）
        if not self.config.offline_mode:
            self._preload_models()

    def _preload_models(self):
        """预下载必要的模型文件，避免运行时下载超时"""
        try:
            if self.config.partition_strategy == "hi_res":
                self.logger.info("预下载hi_res策略所需的模型文件...")
                self._download_yolo_models()
            elif self.config.partition_strategy == "fast":
                self.logger.info("使用fast策略，跳过模型预下载")
        except Exception as e:
            self.logger.warning(f"模型预下载失败: {e}")

    def _download_yolo_models(self):
        """下载YOLO模型文件"""
        try:
            from huggingface_hub import hf_hub_download
            import subprocess

            # 模型文件列表
            model_files = [
                {
                    "repo_id": "unstructuredio/yolo_x_layout",
                    "filename": "yolox_l0.05.onnx",
                    "local_dir": "~/.cache/unstructured",
                }
            ]

            for model in model_files:
                try:
                    self.logger.info(f"正在下载模型: {model['filename']}")
                    hf_hub_download(
                        repo_id=model["repo_id"],
                        filename=model["filename"],
                        cache_dir=model["local_dir"],
                        resume_download=True,
                        timeout=self.config.download_timeout,
                    )
                    self.logger.info(f"模型下载完成: {model['filename']}")
                except Exception as e:
                    self.logger.warning(f"模型 {model['filename']} 下载失败: {e}")

        except ImportError:
            self.logger.warning("未安装huggingface_hub，跳过模型预下载")
        except Exception as e:
            self.logger.error(f"模型预下载过程出错: {e}")

    def _build_partition_options(self) -> Dict[str, Any]:
        """构建Unstructured分区选项，使用最小参数避免OCR触发"""
        options = {
            "strategy": self.config.partition_strategy,
        }

        # 只在必要时添加languages参数
        if self.config.languages and self.config.languages != ["auto"]:
            options["languages"] = self.config.languages

        return options

    def process_document(self, file_path: str) -> DocumentContent:
        """
        统一的文档处理接口，支持LLM多模态增强，完全避免OCR

        Args:
            file_path: 文档文件路径

        Returns:
            DocumentContent: 包含提取内容和元信息的结构化对象
        """
        try:
            # 1. 文件存在性检查
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文档文件不存在: {file_path}")

            # 2. 文件类型检测，决定处理策略
            file_extension = os.path.splitext(file_path)[1].lower()
            processing_strategy = self._determine_processing_strategy(file_extension)

            self.logger.info(
                f"文件类型: {file_extension}, 处理策略: {processing_strategy}"
            )

            # 3. 根据文件类型采用不同的处理方式
            if processing_strategy == "pure_text":
                # 纯文本文件：使用回退处理，完全避免unstructured的OCR
                doc_content = self._process_pure_text_file(file_path)
            elif processing_strategy == "structured_text":
                # 结构化文本：使用unstructured，但避免OCR相关参数
                if not UNSTRUCTURED_AVAILABLE:
                    self.logger.warning("unstructured库未安装，使用回退处理方案")
                    doc_content = self._fallback_process_document(file_path)
                else:
                    # 使用最简化的unstructured参数，避免触发OCR
                    doc_content = self._process_with_minimal_unstructured(file_path)
            else:
                # Office文档和PDF：使用unstructured，依赖fast策略减少OCR
                if not UNSTRUCTURED_AVAILABLE:
                    self.logger.warning("unstructured库未安装，使用回退处理方案")
                    doc_content = self._fallback_process_document(file_path)
                else:
                    elements = partition(filename=file_path, **self.partition_options)
                    doc_content = self._extract_structured_content(elements, file_path)

            # 4. 如果启用多模态，额外处理图片内容
            if self.config.enable_multimodal and self.multimodal_processor:
                try:
                    self.logger.info(f"开始LLM多模态处理: {file_path}")
                    image_contents = (
                        self.multimodal_processor.extract_and_process_images(file_path)
                    )

                    # 将图片内容合并到文档内容中
                    if image_contents:
                        doc_content = self._merge_image_content(
                            doc_content, image_contents
                        )
                        self.logger.info(
                            f"LLM多模态处理完成，提取到 {len(image_contents)} 张图片"
                        )
                    else:
                        self.logger.info("未发现图片内容")

                except Exception as e:
                    self.logger.warning(f"LLM多模态处理失败，使用文本处理结果: {e}")

            return doc_content

        except Exception as e:
            self.logger.error(f"文档处理失败 {file_path}: {e}")
            return self._handle_processing_error(file_path, e)

    def _determine_processing_strategy(self, file_extension: str) -> str:
        """根据文件扩展名确定处理策略，避免OCR"""
        # 纯文本文件：完全避免unstructured，使用简单文本读取
        pure_text_extensions = [
            ".txt",
            ".md",
            ".csv",
            ".json",
            ".log",
            ".xml",
            ".html",
            ".htm",
        ]

        # 结构化文本：可以使用unstructured但不触发OCR
        structured_text_extensions = [".rtf", ".msg"]

        if file_extension in pure_text_extensions:
            return "pure_text"
        elif file_extension in structured_text_extensions:
            return "structured_text"
        else:
            # Office文档、PDF、图片等：使用unstructured但控制OCR使用
            return "complex_document"

    def _process_pure_text_file(self, file_path: str) -> DocumentContent:
        """处理纯文本文件，完全避免OCR和unstructured"""
        self.logger.info(f"使用纯文本处理模式: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            return DocumentContent(
                sections=[
                    {
                        "type": "text",
                        "content": content,
                        "metadata": {"processing_mode": "pure_text", "no_ocr": True},
                    }
                ],
                metadata={
                    "file_path": file_path,
                    "file_type": self._detect_file_type(file_path),
                    "processing_timestamp": datetime.now().isoformat(),
                    "processing_mode": "pure_text",
                    "no_unstructured_ocr": True,
                    "element_count": 1,
                },
                full_text=content,
            )
        except UnicodeDecodeError:
            # 尝试其他编码
            try:
                with open(file_path, "r", encoding="gbk") as f:
                    content = f.read()

                return DocumentContent(
                    sections=[
                        {
                            "type": "text",
                            "content": content,
                            "metadata": {
                                "processing_mode": "pure_text_gbk",
                                "no_ocr": True,
                            },
                        }
                    ],
                    metadata={
                        "file_path": file_path,
                        "file_type": self._detect_file_type(file_path),
                        "processing_timestamp": datetime.now().isoformat(),
                        "processing_mode": "pure_text_gbk",
                        "no_unstructured_ocr": True,
                        "element_count": 1,
                        "encoding": "gbk",
                    },
                    full_text=content,
                )
            except Exception as e:
                return self._handle_processing_error(file_path, e)

    def _process_with_minimal_unstructured(self, file_path: str) -> DocumentContent:
        """使用最小化参数的unstructured处理，避免OCR触发"""
        self.logger.info(f"使用最小化unstructured处理: {file_path}")

        try:
            # 只传递最基本的参数，避免触发OCR
            elements = partition(
                filename=file_path,
                strategy="fast",  # 使用fast策略，减少OCR使用
                # 注意：故意不传递languages、model_name等可能触发OCR的参数
            )

            return self._extract_structured_content(elements, file_path)

        except Exception as e:
            self.logger.warning(f"最小化unstructured处理失败，回退到完整处理: {e}")
            # 回退到完整的unstructured处理
            elements = partition(filename=file_path, **self.partition_options)
            return self._extract_structured_content(elements, file_path)

    def _extract_structured_content(self, elements, file_path: str) -> DocumentContent:
        """提取结构化内容"""
        content_sections = []
        metadata = {
            "file_path": file_path,
            "file_type": self._detect_file_type(file_path),
            "processing_timestamp": datetime.now().isoformat(),
            "element_count": len(elements) if elements else 0,
        }

        if elements:
            for element in elements:
                if hasattr(element, "text") and element.text.strip():
                    content_sections.append(
                        {
                            "type": getattr(element, "category", "text"),
                            "content": element.text.strip(),
                            "metadata": getattr(element, "metadata", {}),
                        }
                    )

        return DocumentContent(
            sections=content_sections,
            metadata=metadata,
            full_text="\n".join([s["content"] for s in content_sections]),
        )

    @staticmethod
    def _safe_get_metadata_value(metadata, key, default=None):
        """安全获取metadata值，支持字典和ElementMetadata对象"""
        if hasattr(metadata, "get"):
            # 如果是字典类型
            return metadata.get(key, default)
        elif hasattr(metadata, key):
            # 如果是ElementMetadata对象类型
            return getattr(metadata, key, default)
        else:
            # 如果metadata不是对象也不是字典
            return default

    @staticmethod
    def _safe_get_section_metadata(section, key, default=None):
        """安全获取section的metadata值"""
        metadata = section.get("metadata", {})
        return UnifiedDocumentProcessor._safe_get_metadata_value(metadata, key, default)

    def _find_optimal_insert_position(self, sections):
        """找到最佳的图片内容插入位置"""
        if not sections:
            return 0

        # 在文档的前1/3位置插入，或在第一个非文本section之后
        optimal_position = min(len(sections) // 3, len(sections) - 1)

        # 如果前面有标题或目录，跳过
        for i, section in enumerate(sections[:optimal_position]):
            content = section.get("content", "").lower()
            title_keywords = [
                "目录",
                "contents",
                "目录",
                "index",
                "索引",
                "summary",
                "概述",
                "摘要",
            ]

            if any(keyword in content for keyword in title_keywords):
                optimal_position = i + 1
                break

        # 确保不在文档末尾
        if optimal_position >= len(sections):
            optimal_position = len(sections) - 1

        return max(1, optimal_position)  # 确保至少在第2个位置

    def _rebuild_full_text(self, doc_content):
        """重新构建完整文本，确保LLM内容不在最后面"""
        all_sections_text = []

        for i, section in enumerate(doc_content.sections):
            if section.get("content"):
                # 为图片内容添加标识
                if section.get("type") == "image_content" and section.get(
                    "metadata", {}
                ).get("llm_processed"):
                    # LLM增强的图片内容，特殊处理
                    content = section["content"]
                    # 移除重复的标题（因为content已经包含标题）
                    if content.startswith("\n## 图片内容识别结果"):
                        content = content.replace(
                            "\n## 图片内容识别结果", "", 1
                        ).strip()

                    all_sections_text.append(
                        f"\n## 图片内容识别 (LLM增强)\n\n{content}"
                    )
                else:
                    # 普通内容
                    all_sections_text.append(section["content"])

        # 如果有LLM内容，确保它们不在文档末尾
        llm_sections = [
            s
            for s in doc_content.sections
            if s.get("metadata", {}).get("llm_processed")
        ]
        if llm_sections and len(all_sections_text) > 1:
            # 将最后的LLM内容移到中间位置
            non_llm_sections = [
                text
                for text in all_sections_text
                if not text.startswith("\n## 图片内容识别")
            ]

            if non_llm_sections:
                # 在非LLM内容的1/3处插入LLM内容
                insert_pos = len(non_llm_sections) // 3
                remaining_llm = [
                    text
                    for text in all_sections_text
                    if text.startswith("\n## 图片内容识别")
                ]

                final_sections = (
                    non_llm_sections[:insert_pos]
                    + remaining_llm
                    + non_llm_sections[insert_pos:]
                )
                doc_content.full_text = "\n\n".join(final_sections)
            else:
                doc_content.full_text = "\n\n".join(all_sections_text)
        else:
            doc_content.full_text = "\n\n".join(all_sections_text)

    def _merge_image_content(
        self, doc_content: DocumentContent, image_contents: List[ImageContentInfo]
    ) -> DocumentContent:
        """合并图片内容到文档内容中，通过LLM增强识别结果"""
        try:
            # 检查是否需要替换现有的图片内容
            image_sections_to_replace = []
            original_image_contents = []  # 收集原始图片内容用于替换

            self.logger.info(
                f"检查 {len(doc_content.sections)} 个sections以识别图片内容"
            )

            for i, section in enumerate(doc_content.sections):
                section_type = section.get("type", "")
                section_content = section.get("content", "")
                section_metadata = section.get("metadata", {})

                # 更宽松的识别条件
                is_image_section = False

                # 1. 检查section类型
                if section_type in [
                    "image",
                    "figure",
                    "picture",
                    "image_content",
                    "Image",
                    "Figure",
                ]:
                    is_image_section = True
                    self.logger.debug(f"Section {i}: 类型匹配 - {section_type}")

                # 2. 检查content是否包含图片相关关键词
                elif any(
                    keyword in section_content.lower()
                    for keyword in [
                        "image",
                        "图片",
                        "figure",
                        "图",
                        "picture",
                        "photo",
                        "photo",
                        "截图",
                        "图表",
                    ]
                ):
                    is_image_section = True
                    self.logger.debug(
                        f"Section {i}: 内容关键词匹配 - {section_content[:50]}..."
                    )

                # 3. 检查metadata
                elif (
                    self._safe_get_section_metadata(section, "element_type")
                    in ["Image", "image"]
                    or self._safe_get_section_metadata(section, "category")
                    in ["Image", "image"]
                    or self._safe_get_section_metadata(section, "type")
                    in ["Image", "image"]
                ):
                    is_image_section = True
                    self.logger.debug(f"Section {i}: metadata匹配")

                # 4. 特别检查：如果是第一个非文本section，也认为是图片内容
                elif (
                    i > 0
                    and section_type == "text"
                    and len(section_content) < 100
                    and not any(char.isdigit() for char in section_content)  # 没有数字
                    and section_content.strip()  # 非空
                    and doc_content.sections[max(0, i - 1)].get("type") == "text"
                ):  # 前面也是文本
                    is_image_section = True
                    self.logger.debug(f"Section {i}: 特殊模式识别 - 可能是图片内容")

                # 6. 如果content很长且包含图片相关描述
                elif len(section_content) < 200 and any(
                    desc in section_content.lower()
                    for desc in [
                        "包含",
                        "contains",
                        "显示",
                        "shows",
                        "描述",
                        "describes",
                        "截图",
                        "screenshot",
                    ]
                ):
                    is_image_section = True
                    self.logger.debug(f"Section {i}: 描述性内容匹配")

                # 记录识别结果
                if is_image_section:
                    image_sections_to_replace.append(i)
                    original_image_contents.append(section)
                    self.logger.info(f"识别为图片内容: Section {i} - {section_type}")

            self.logger.info(
                f"总共识别到 {len(image_sections_to_replace)} 个图片内容sections"
            )

            # 如果有LLM图片内容，不管是否识别到原始图片内容，都要进行替换或插入
            if image_contents:
                self.logger.info(
                    f"发现 {len(image_sections_to_replace)} 个低质量图片内容，将用LLM结果替换"
                )

                # 创建新的sections列表，逐步替换
                new_sections = []
                current_image_index = 0

                for i, section in enumerate(doc_content.sections):
                    if i in image_sections_to_replace and current_image_index < len(
                        image_contents
                    ):
                        # 替换为LLM识别的高质量内容
                        img_info = image_contents[current_image_index]
                        if img_info.processed_content:
                            img_section = {
                                "type": "image_content",
                                "content": f"\n## 图片内容识别结果\n\n{img_info.processed_content}",
                                "metadata": {
                                    "image_id": img_info.image_id,
                                    "alt_text": img_info.alt_text,
                                    "extraction_type": img_info.extraction_type,
                                    "source_path": img_info.source_path,
                                    "llm_processed": True,
                                    "replaced_original": True,
                                },
                            }
                            new_sections.append(img_section)
                            current_image_index += 1
                        else:
                            # 如果LLM处理失败，保留原始内容
                            new_sections.append(section)
                    else:
                        # 非图片内容或多余的图片内容
                        new_sections.append(section)

                # 添加剩余的LLM识别结果（如果没有足够的原始图片内容可替换）
                while current_image_index < len(image_contents):
                    img_info = image_contents[current_image_index]
                    if img_info.processed_content:
                        img_section = {
                            "type": "image_content",
                            "content": f"\n## 图片内容识别结果 (ID: {img_info.image_id})\n\n{img_info.processed_content}",
                            "metadata": {
                                "image_id": img_info.image_id,
                                "alt_text": img_info.alt_text,
                                "extraction_type": img_info.extraction_type,
                                "source_path": img_info.source_path,
                                "llm_processed": True,
                                "replaced_original": False,
                            },
                        }
                        new_sections.append(img_section)
                    current_image_index += 1

                doc_content.sections = new_sections

            else:
                # 如果没有找到需要替换的内容，智能插入到合适位置而不是末尾
                self.logger.info("未找到可替换的图片内容，将智能插入LLM结果")

                # 尝试在文档中间合适位置插入第一个LLM图片内容
                insert_position = self._find_optimal_insert_position(
                    doc_content.sections
                )

                for img_info in image_contents:
                    if img_info.processed_content:
                        img_section = {
                            "type": "image_content",
                            "content": f"\n## 图片内容识别结果\n\n{img_info.processed_content}",
                            "metadata": {
                                "image_id": img_info.image_id,
                                "alt_text": img_info.alt_text,
                                "extraction_type": img_info.extraction_type,
                                "source_path": img_info.source_path,
                                "llm_processed": True,
                                "inserted_position": insert_position,
                            },
                        }

                        # 插入到合适位置而不是末尾
                        if insert_position < len(doc_content.sections):
                            doc_content.sections.insert(insert_position, img_section)
                        else:
                            # 如果插入位置超出范围，添加到倒数第二个位置
                            if len(doc_content.sections) > 1:
                                doc_content.sections.insert(-1, img_section)
                            else:
                                # 如果只有1个section，插入到第一个位置
                                doc_content.sections.insert(0, img_section)

                        insert_position += 1  # 下一个插入位置

                self.logger.info(f"已智能插入 {len(image_contents)} 个LLM图片内容")

            # 重新构建完整文本，保持正确的顺序
            self._rebuild_full_text(doc_content)

            # 更新元数据
            doc_content.metadata["image_count"] = len(image_contents)
            doc_content.metadata["llm_multimodal_processed"] = True
            doc_content.metadata["multimodal_provider"] = (
                self.config.multimodal_config.vision_model_provider
                if self.config.multimodal_config
                else None
            )

            if image_sections_to_replace:
                # 替换模式
                doc_content.metadata["replaced_image_sections"] = len(
                    image_sections_to_replace
                )
                doc_content.metadata["original_image_sections_kept"] = len(
                    image_contents
                ) - min(len(image_contents), len(image_sections_to_replace))
                doc_content.metadata["processing_mode"] = "replace"
            else:
                # 插入模式
                doc_content.metadata["replaced_image_sections"] = 0
                doc_content.metadata["inserted_image_sections"] = len(image_contents)
                doc_content.metadata["processing_mode"] = "insert"

            # 清理临时文件
            if hasattr(self, "multimodal_processor"):
                self.multimodal_processor.cleanup_temporary_files(image_contents)

            return doc_content

        except Exception as e:
            self.logger.error(f"合并图片内容失败: {e}")
            return doc_content

    def _detect_file_type(self, file_path: str) -> str:
        """检测文件类型"""
        mime_type, _ = mimetypes.guess_type(file_path)
        return mime_type or "unknown"

    def _fallback_process_document(self, file_path: str) -> DocumentContent:
        """回退处理方案（当unstructured不可用时）"""
        self.logger.info(f"使用回退方案处理: {file_path}")

        try:
            _, extension = os.path.splitext(file_path.lower())

            if extension == ".txt":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".md":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".csv":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".json":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            else:
                content = f"不支持的文件格式: {extension}\n请安装unstructured库以支持更多格式。"

            return DocumentContent(
                sections=[
                    {"type": "text", "content": content, "metadata": {"fallback": True}}
                ],
                metadata={
                    "file_path": file_path,
                    "file_type": extension,
                    "processing_timestamp": datetime.now().isoformat(),
                    "fallback_method": True,
                },
                full_text=content,
            )

        except Exception as e:
            return self._handle_processing_error(file_path, e)

    def _handle_processing_error(
        self, file_path: str, error: Exception
    ) -> DocumentContent:
        """处理处理错误"""
        error_content = f"文档处理失败: {str(error)}"

        return DocumentContent(
            sections=[
                {
                    "type": "error",
                    "content": error_content,
                    "metadata": {"error_type": type(error).__name__},
                }
            ],
            metadata={
                "file_path": file_path,
                "processing_error": str(error),
                "error_type": type(error).__name__,
                "processing_timestamp": datetime.now().isoformat(),
            },
            full_text=error_content,
        )

    def get_supported_formats(self) -> List[str]:
        """获取支持的文档格式列表"""
        if UNSTRUCTURED_AVAILABLE:
            return [
                ".pdf",
                ".docx",
                ".doc",
                ".txt",
                ".md",
                ".rtf",
                ".html",
                ".htm",
                ".xml",
                ".json",
                ".csv",
                ".pptx",
                ".ppt",
                ".xlsx",
                ".xls",
                ".odt",
                ".epub",
                ".mobi",
                ".log",
                ".msg",
            ]
        else:
            return [".txt", ".md", ".csv", ".json"]


# ==================== Excel处理模块 ====================


class ExcelLinkProcessor:
    """Excel超链接处理器"""

    def __init__(self, document_processor: UnifiedDocumentProcessor):
        self.document_processor = document_processor
        self.logger = logging.getLogger(__name__)

    def find_hyperlinks(self, sheet) -> List[HyperlinkInfo]:
        """查找Excel中的超链接"""
        links = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    links.append(
                        HyperlinkInfo(
                            cell=cell,
                            target=cell.hyperlink.target,
                            display_text=cell.value,
                        )
                    )

        self.logger.info(f"找到 {len(links)} 个超链接")
        return links

    def resolve_path(self, link_target: str, base_dir: str) -> str:
        """
        解析链接路径，支持相对和绝对路径

        Args:
            link_target: 链接目标路径
            base_dir: Excel文件所在目录

        Returns:
            str: 解析后的绝对路径
        """
        try:
            if os.path.isabs(link_target):
                return link_target

            # 处理相对路径
            resolved_path = os.path.join(base_dir, link_target)
            return os.path.normpath(resolved_path)

        except Exception as e:
            self.logger.error(f"路径解析失败 {link_target}: {e}")
            return link_target


# ==================== 内容格式化模块 ====================


class ContentFormatter:
    """内容格式化器"""

    def __init__(self, config: DocumentProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)

    def format_content(self, doc_content: DocumentContent) -> str:
        """格式化文档内容"""
        if self.config.output_format == "markdown":
            return self._to_markdown(doc_content)
        elif self.config.output_format == "plain_text":
            return doc_content.full_text
        elif self.config.output_format == "json":
            import json

            return json.dumps(doc_content.__dict__, ensure_ascii=False, indent=2)
        else:
            return self._to_markdown(doc_content)  # 默认返回markdown

    def _to_markdown(self, doc_content: DocumentContent) -> str:
        """转换为Markdown格式"""
        markdown_parts = []

        # 添加文件信息头部（如果启用）
        if self.config.include_metadata:
            markdown_parts.append("```metadata")
            markdown_parts.append(
                f"文件: {doc_content.metadata.get('file_path', '未知')}"
            )
            markdown_parts.append(
                f"类型: {doc_content.metadata.get('file_type', '未知')}"
            )
            markdown_parts.append(
                f"处理时间: {doc_content.metadata.get('processing_timestamp', '未知')}"
            )
            if "element_count" in doc_content.metadata:
                markdown_parts.append(
                    f"元素数量: {doc_content.metadata['element_count']}"
                )
            if doc_content.metadata.get("fallback_method"):
                markdown_parts.append("注意: 使用回退方案处理")
            if doc_content.metadata.get("llm_multimodal_processed"):
                markdown_parts.append(
                    f"LLM多模态处理: OK ({doc_content.metadata.get('multimodal_provider', 'unknown')})"
                )
                if "image_count" in doc_content.metadata:
                    markdown_parts.append(
                        f"图片数量: {doc_content.metadata['image_count']}"
                    )
            markdown_parts.append("```")
            markdown_parts.append("")

        # 添加内容段落
        for section in doc_content.sections:
            content = section["content"]

            # 内容长度限制
            if len(content) > self.config.max_content_length:
                content = (
                    content[: self.config.max_content_length] + "...\n\n[内容已截断]"
                )

            # 添加类型标识（如果内容不为空）
            if content.strip():
                section_type = section.get("type", "text")
                section_metadata = section.get("metadata", {})

                def _safe_check(metadata, key):
                    """安全检查metadata中的布尔值"""
                    if hasattr(metadata, "get"):
                        return metadata.get(key, False)
                    elif hasattr(metadata, key):
                        return getattr(metadata, key, False)
                    else:
                        return False

                if section_type == "image_content":
                    if _safe_check(section_metadata, "llm_processed"):
                        if _safe_check(section_metadata, "replaced_original"):
                            markdown_parts.append(
                                f"### 图片内容识别 (LLM增强 - 已替换)\n\n{content}\n"
                            )
                        else:
                            markdown_parts.append(
                                f"### 图片内容识别 (LLM增强)\n\n{content}\n"
                            )
                    else:
                        markdown_parts.append(f"### 图片内容识别\n\n{content}\n")
                else:
                    markdown_parts.append(f"```\n{content}\n```")

        if not markdown_parts:
            markdown_parts.append("```\n[无内容]\n```")

        return "\n".join(markdown_parts)


# ==================== 主要接口类 ====================


class EnhancedExcelProcessorLLM:
    """增强的Excel处理器 - LLM多模态版"""

    def __init__(self, config: Optional[DocumentProcessingConfig] = None):
        self.config = config or DocumentProcessingConfig()

        self.document_processor = UnifiedDocumentProcessor(self.config)
        self.content_formatter = ContentFormatter(self.config)
        self.excel_processor = ExcelLinkProcessor(self.document_processor)

        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler("excel_processor_llm.log", encoding="utf-8"),
            ],
        )
        self.logger = logging.getLogger(__name__)

        self.logger.info("LLM增强版Excel处理器初始化完成")

    def process_excel_file(self, excel_path: str) -> Dict[str, Any]:
        """
        处理Excel文件中的链接文档，支持LLM多模态增强

        Args:
            excel_path: Excel文件路径

        Returns:
            Dict: 处理结果
        """
        results = {
            "file_path": excel_path,
            "processed_at": datetime.now().isoformat(),
            "total_links": 0,
            "successful": 0,
            "failed": 0,
            "errors": [],
            "multimodal_processed": False,
        }

        workbook = None
        try:
            # 1. 加载Excel文件
            self.logger.info(f"加载Excel文件: {excel_path}")
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active

            # 2. 查找超链接
            links = self.excel_processor.find_hyperlinks(sheet)
            results["total_links"] = len(links)

            if not links:
                self.logger.info(f"Excel文件 '{excel_path}' 中未找到超链接")
                return results

            # 4. 设置目标列
            first_link_col = links[0].cell.column
            content_col = first_link_col + 1

            sheet.insert_cols(content_col)

            # 设置标题
            header_cell = sheet.cell(row=1, column=content_col)
            header_cell.value = "链接文档内容 (LLM增强)"
            header_cell.font = Font(bold=True)
            self.logger.info(f"在第 {get_column_letter(content_col)} 列插入内容列")

            # 5. 处理每个链接
            for link_info in links:
                try:
                    # 解析文件路径
                    base_dir = os.path.dirname(os.path.abspath(excel_path))
                    full_path = self.excel_processor.resolve_path(
                        link_info.target, base_dir
                    )

                    self.logger.info(f"处理链接: {link_info.target} -> {full_path}")

                    # 提取文档内容（包含LLM多模态处理）
                    doc_content = self.document_processor.process_document(full_path)

                    # 检查是否进行了多模态处理
                    if doc_content.metadata.get("llm_multimodal_processed"):
                        results["multimodal_processed"] = True

                    # 格式化内容
                    formatted_content = self.content_formatter.format_content(
                        doc_content
                    )

                    # 更新Excel单元格
                    content_cell = sheet.cell(
                        row=link_info.cell.row, column=content_col
                    )
                    content_cell.value = formatted_content

                    results["successful"] += 1
                    self.logger.info(f"处理完成: {link_info.target}")

                except Exception as e:
                    error_msg = f"处理链接 '{link_info.target}' 失败: {str(e)}"
                    results["errors"].append(error_msg)
                    results["failed"] += 1
                    self.logger.error(error_msg)

                    # 如果不继续处理错误，则停止
                    if not self.config.continue_on_error:
                        break

            # 6. 保存文件
            workbook.save(excel_path)
            self.logger.info(f"Excel文件已更新: {excel_path}")

        except Exception as e:
            error_msg = f"Excel文件处理失败: {str(e)}"
            results["errors"].append(error_msg)
            self.logger.error(error_msg)

        finally:
            if workbook:
                workbook.close()

        return results

    def process_multiple_files(
        self, file_paths: List[str], max_workers: int = None
    ) -> Dict[str, Dict[str, Any]]:
        """批量处理多个Excel文件"""
        if not max_workers:
            max_workers = (
                self.config.max_workers if hasattr(self.config, "max_workers") else 4
            )

        if not self.config.enable_parallel_processing:
            max_workers = 1

        results = {}
        self.logger.info(
            f"开始批量处理 {len(file_paths)} 个文件，使用 {max_workers} 个线程"
        )

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交任务
            future_to_file = {
                executor.submit(self.process_excel_file, file_path): file_path
                for file_path in file_paths
            }

            # 收集结果
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    results[file_path] = future.result()
                except Exception as e:
                    results[file_path] = {
                        "file_path": file_path,
                        "error": str(e),
                        "successful": 0,
                        "failed": 0,
                        "processed_at": datetime.now().isoformat(),
                    }
                    self.logger.error(f"批量处理文件失败 {file_path}: {e}")

        self.logger.info(
            f"批量处理完成，成功: {sum(r.get('successful', 0) for r in results.values())}"
        )
        return results

    def get_supported_formats(self) -> List[str]:
        """获取支持的文档格式列表"""
        return self.document_processor.get_supported_formats()


# ==================== 便捷接口 ====================


def create_llm_processor(
    config: Optional[DocumentProcessingConfig] = None,
) -> EnhancedExcelProcessorLLM:
    """创建LLM增强版Excel处理器的便捷函数"""
    return EnhancedExcelProcessorLLM(config)


def process_excel_links_llm(
    excel_path: str,
    config: Optional[DocumentProcessingConfig] = None,
) -> Dict[str, Any]:
    """
    处理Excel文件中链接的LLM增强版便捷函数

    Args:
        excel_path: Excel文件路径
        config: 处理配置

    Returns:
        Dict: 处理结果
    """
    processor = create_llm_processor(config)
    return processor.process_excel_file(excel_path)


# ==================== 使用示例 ====================

if __name__ == "__main__":
    print("Excel链接内容提取器 v2.0.0 - LLM多模态增强版")
    print("采用多模态大模型进行高精度图片内容识别和解析")
    print()

    # 1. 创建多模态配置
    multimodal_config = MultimodalConfig(
        vision_model_provider="qwen",  # 使用qwen提供商
        vision_model_name="qwen-vl-plus",  # 通义千问视觉模型
        api_key=os.environ.get("QWEN_V"),  # 需要设置DashScope API密钥
        api_base="https://dashscope.aliyuncs.com/compatible-mode/v1",  # OpenAI兼容端点
        extract_embedded_images=True,
        vision_prompt="请详细识别和描述这张图片中的所有文本内容，包括标题、正文、表格、图表等，保持原有的格式和结构。如果是表格，请用markdown表格格式输出。",
        max_tokens_per_image=2000,
        enable_caching=True,
    )

    # 2. 创建处理器配置
    config = DocumentProcessingConfig(
        output_format="markdown",
        enable_parallel_processing=False,  # 单文件处理设为False
        enable_multimodal=True,  # 启用LLM多模态功能
        multimodal_config=multimodal_config,
    )

    # 3. 创建处理器
    processor = EnhancedExcelProcessorLLM(config)

    # 5. 示例用法（需要修改为实际的Excel文件路径）
    excel_file = "C:\\Users\\Admin\\Desktop\\text\\任务管理.xlsx"

    print(f"\n开始处理Excel文件: {excel_file}")
    print("=" * 60)

    print("开始实际处理...")
    result = processor.process_excel_file(excel_file)
    print(f"处理结果:")
    print(f"成功处理: {result['successful']} 个链接")
    print(f"处理失败: {result['failed']} 个链接")
    print(f"LLM多模态处理: {'是' if result['multimodal_processed'] else '否'}")

    if result["errors"]:
        print(f"错误信息:")
        for error in result["errors"]:
            print(f"    - {error}")

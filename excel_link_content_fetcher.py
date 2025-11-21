# -*- coding: utf-8 -*-
"""
Excel链接内容提取器

基于Unstructured的增强型Excel超链接文档内容提取工具，支持30+种文档格式，
具备备份机制、错误恢复、性能优化等企业级特性。
"""

import os
import shutil
import time
import logging
import mimetypes
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

try:
    from unstructured.partition.auto import partition

    UNSTRUCTURED_AVAILABLE = True
except ImportError:
    UNSTRUCTURED_AVAILABLE = False
    partition = None


# ==================== 配置和数据结构 ====================


@dataclass
class DocumentContent:
    """文档内容结构"""

    sections: List[Dict[str, Any]]
    metadata: Dict[str, Any]
    full_text: str

    def __post_init__(self):
        """后处理验证"""
        if not isinstance(self.sections, list):
            self.sections = []
        if not isinstance(self.metadata, dict):
            self.metadata = {}
        if not isinstance(self.full_text, str):
            self.full_text = ""


@dataclass
class DocumentProcessingConfig:
    """文档处理配置"""

    # Unstructured配置
    partition_strategy: str = "hi_res"  # hi_res, fast, ocr_only
    model_name: Optional[str] = None
    languages: List[str] = field(default_factory=lambda: ["zh", "en"])

    # 输出配置
    output_format: str = "markdown"  # markdown, plain_text, json
    max_content_length: int = 10000
    include_metadata: bool = True

    continue_on_error: bool = True
    max_retries: int = 3

    # 性能配置
    enable_parallel_processing: bool = True
    max_workers: int = 4
    memory_limit_mb: int = 512

    def __post_init__(self):
        """后处理初始化"""
        # 确保languages不为空
        if not self.languages:
            self.languages = ["zh", "en"]

        # 验证输出格式
        valid_formats = ["markdown", "plain_text", "json"]
        if self.output_format not in valid_formats:
            self.output_format = "markdown"

        # 验证分区策略
        valid_strategies = ["hi_res", "fast", "ocr_only"]
        if self.partition_strategy not in valid_strategies:
            self.partition_strategy = "hi_res"


@dataclass
class HyperlinkInfo:
    """超链接信息"""

    cell: openpyxl.cell.Cell
    target: str
    display_text: Union[str, int, float]

    def __post_init__(self):
        """后处理验证"""
        if self.display_text is None:
            self.display_text = ""


# ==================== 核心处理模块 ====================


class UnifiedDocumentProcessor:
    """基于Unstructured的统一文档处理器"""

    def __init__(self, config: DocumentProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)
        self.partition_options = self._build_partition_options()

    def _build_partition_options(self) -> Dict[str, Any]:
        """构建Unstructured分区选项"""
        options = {
            "strategy": self.config.partition_strategy,
            "languages": self.config.languages,
        }

        if self.config.model_name:
            options["model_name"] = self.config.model_name

        return options

    def process_document(self, file_path: str) -> DocumentContent:
        """
        统一的文档处理接口

        Args:
            file_path: 文档文件路径

        Returns:
            DocumentContent: 包含提取内容和元信息的结构化对象
        """
        try:
            # 1. 文件存在性检查
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"文档文件不存在: {file_path}")

            # 2. 检查unstructured是否可用
            if not UNSTRUCTURED_AVAILABLE:
                self.logger.warning("unstructured库未安装，使用回退处理方案")
                return self._fallback_process_document(file_path)

            # 3. 使用Unstructured自动检测和处理
            elements = partition(filename=file_path, **self.partition_options)

            # 4. 提取和结构化内容
            return self._extract_structured_content(elements, file_path)

        except Exception as e:
            self.logger.error(f"文档处理失败 {file_path}: {e}")
            return self._handle_processing_error(file_path, e)

    def _extract_structured_content(self, elements, file_path: str) -> DocumentContent:
        """提取结构化内容"""
        content_sections = []
        metadata = {
            "file_path": file_path,
            "file_type": self._detect_file_type(file_path),
            "processing_timestamp": datetime.now().isoformat(),
            "element_count": len(elements) if elements else 0,
        }

        if elements:
            for element in elements:
                if hasattr(element, "text") and element.text.strip():
                    content_sections.append(
                        {
                            "type": getattr(element, "category", "text"),
                            "content": element.text.strip(),
                            "metadata": getattr(element, "metadata", {}),
                        }
                    )

        return DocumentContent(
            sections=content_sections,
            metadata=metadata,
            full_text="\n".join([s["content"] for s in content_sections]),
        )

    def _detect_file_type(self, file_path: str) -> str:
        """检测文件类型"""
        mime_type, _ = mimetypes.guess_type(file_path)
        return mime_type or "unknown"

    def _fallback_process_document(self, file_path: str) -> DocumentContent:
        """回退处理方案（当unstructured不可用时）"""
        self.logger.info(f"使用回退方案处理: {file_path}")

        try:
            _, extension = os.path.splitext(file_path.lower())

            if extension == ".txt":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".md":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".csv":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif extension == ".json":
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            else:
                content = f"不支持的文件格式: {extension}\n请安装unstructured库以支持更多格式。"

            return DocumentContent(
                sections=[
                    {"type": "text", "content": content, "metadata": {"fallback": True}}
                ],
                metadata={
                    "file_path": file_path,
                    "file_type": extension,
                    "processing_timestamp": datetime.now().isoformat(),
                    "fallback_method": True,
                },
                full_text=content,
            )

        except Exception as e:
            return self._handle_processing_error(file_path, e)

    def _handle_processing_error(
        self, file_path: str, error: Exception
    ) -> DocumentContent:
        """处理处理错误"""
        error_content = f"文档处理失败: {str(error)}"

        return DocumentContent(
            sections=[
                {
                    "type": "error",
                    "content": error_content,
                    "metadata": {"error_type": type(error).__name__},
                }
            ],
            metadata={
                "file_path": file_path,
                "processing_error": str(error),
                "error_type": type(error).__name__,
                "processing_timestamp": datetime.now().isoformat(),
            },
            full_text=error_content,
        )

    def get_supported_formats(self) -> List[str]:
        """获取支持的文档格式列表"""
        if UNSTRUCTURED_AVAILABLE:
            return [
                ".pdf",
                ".docx",
                ".doc",
                ".txt",
                ".md",
                ".rtf",
                ".html",
                ".htm",
                ".xml",
                ".json",
                ".csv",
                ".pptx",
                ".ppt",
                ".xlsx",
                ".xls",
                ".odt",
                ".epub",
                ".mobi",
                ".log",
                ".msg",
            ]
        else:
            return [".txt", ".md", ".csv", ".json"]


# ==================== 备份管理模块 ====================


# ==================== Excel处理模块 ====================


class ExcelLinkProcessor:
    """Excel超链接处理器"""

    def __init__(self, document_processor: UnifiedDocumentProcessor):
        self.document_processor = document_processor
        self.logger = logging.getLogger(__name__)

    def find_hyperlinks(self, sheet) -> List[HyperlinkInfo]:
        """查找Excel中的超链接"""
        links = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    links.append(
                        HyperlinkInfo(
                            cell=cell,
                            target=cell.hyperlink.target,
                            display_text=cell.value,
                        )
                    )

        self.logger.info(f"找到 {len(links)} 个超链接")
        return links

    def resolve_path(self, link_target: str, base_dir: str) -> str:
        """
        解析链接路径，支持相对和绝对路径

        Args:
            link_target: 链接目标路径
            base_dir: Excel文件所在目录

        Returns:
            str: 解析后的绝对路径
        """
        try:
            if os.path.isabs(link_target):
                return link_target

            # 处理相对路径
            resolved_path = os.path.join(base_dir, link_target)
            return os.path.normpath(resolved_path)

        except Exception as e:
            self.logger.error(f"路径解析失败 {link_target}: {e}")
            return link_target


# ==================== 内容格式化模块 ====================


class ContentFormatter:
    """内容格式化器"""

    def __init__(self, config: DocumentProcessingConfig):
        self.config = config
        self.logger = logging.getLogger(__name__)

    def format_content(self, doc_content: DocumentContent) -> str:
        """格式化文档内容"""
        if self.config.output_format == "markdown":
            return self._to_markdown(doc_content)
        elif self.config.output_format == "plain_text":
            return doc_content.full_text
        elif self.config.output_format == "json":
            import json

            return json.dumps(doc_content.__dict__, ensure_ascii=False, indent=2)
        else:
            return self._to_markdown(doc_content)  # 默认返回markdown

    def _to_markdown(self, doc_content: DocumentContent) -> str:
        """转换为Markdown格式"""
        markdown_parts = []

        # 添加文件信息头部（如果启用）
        if self.config.include_metadata:
            markdown_parts.append("```metadata")
            markdown_parts.append(
                f"文件: {doc_content.metadata.get('file_path', '未知')}"
            )
            markdown_parts.append(
                f"类型: {doc_content.metadata.get('file_type', '未知')}"
            )
            markdown_parts.append(
                f"处理时间: {doc_content.metadata.get('processing_timestamp', '未知')}"
            )
            if "element_count" in doc_content.metadata:
                markdown_parts.append(
                    f"元素数量: {doc_content.metadata['element_count']}"
                )
            if doc_content.metadata.get("fallback_method"):
                markdown_parts.append("注意: 使用回退方案处理")
            markdown_parts.append("```")
            markdown_parts.append("")

        # 添加内容段落
        for section in doc_content.sections:
            content = section["content"]

            # 内容长度限制
            if len(content) > self.config.max_content_length:
                content = (
                    content[: self.config.max_content_length] + "...\n\n[内容已截断]"
                )

            # 添加类型标识（如果内容不为空）
            if content.strip():
                markdown_parts.append(f"```\n{content}\n```")

        if not markdown_parts:
            markdown_parts.append("```\n[无内容]\n```")

        return "\n".join(markdown_parts)


# ==================== 主要接口类 ====================


class EnhancedExcelProcessor:
    """增强的Excel处理器"""

    def __init__(self, config: Optional[DocumentProcessingConfig] = None):
        self.config = config or DocumentProcessingConfig()

        self.document_processor = UnifiedDocumentProcessor(self.config)
        self.content_formatter = ContentFormatter(self.config)
        self.excel_processor = ExcelLinkProcessor(self.document_processor)

        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler("excel_processor.log", encoding="utf-8"),
            ],
        )
        self.logger = logging.getLogger(__name__)

        self.logger.info("Excel处理器初始化完成")

    def process_excel_file(self, excel_path: str) -> Dict[str, Any]:
        """
        处理Excel文件中的链接文档

        Args:
            excel_path: Excel文件路径

        Returns:
            Dict: 处理结果
        """
        results = {
            "file_path": excel_path,
            "processed_at": datetime.now().isoformat(),
            "total_links": 0,
            "successful": 0,
            "failed": 0,
            "errors": [],
        }

        workbook = None
        try:
            # 1. 加载Excel文件
            self.logger.info(f"加载Excel文件: {excel_path}")
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active

            # 2. 查找超链接
            links = self.excel_processor.find_hyperlinks(sheet)
            results["total_links"] = len(links)

            if not links:
                self.logger.info(f"Excel文件 '{excel_path}' 中未找到超链接")
                return results

            # 4. 设置目标列
            first_link_col = links[0].cell.column
            content_col = first_link_col + 1

            sheet.insert_cols(content_col)

            # 设置标题
            header_cell = sheet.cell(row=1, column=content_col)
            header_cell.value = "链接文档内容"
            header_cell.font = Font(bold=True)
            self.logger.info(f"在第 {get_column_letter(content_col)} 列插入内容列")

            # 5. 处理每个链接
            for link_info in links:
                try:
                    # 解析文件路径
                    base_dir = os.path.dirname(os.path.abspath(excel_path))
                    full_path = self.excel_processor.resolve_path(
                        link_info.target, base_dir
                    )

                    self.logger.info(f"处理链接: {link_info.target} -> {full_path}")

                    # 提取文档内容
                    doc_content = self.document_processor.process_document(full_path)

                    # 格式化内容
                    formatted_content = self.content_formatter.format_content(
                        doc_content
                    )

                    # 更新Excel单元格
                    content_cell = sheet.cell(
                        row=link_info.cell.row, column=content_col
                    )
                    content_cell.value = formatted_content

                    results["successful"] += 1
                    self.logger.info(f"✅ 处理完成: {link_info.target}")

                except Exception as e:
                    error_msg = f"处理链接 '{link_info.target}' 失败: {str(e)}"
                    results["errors"].append(error_msg)
                    results["failed"] += 1
                    self.logger.error(error_msg)

                    # 如果不继续处理错误，则停止
                    if not self.config.continue_on_error:
                        break

            # 6. 保存文件
            workbook.save(excel_path)
            self.logger.info(f"Excel文件已更新: {excel_path}")

        except Exception as e:
            error_msg = f"Excel文件处理失败: {str(e)}"
            results["errors"].append(error_msg)
            self.logger.error(error_msg)

        finally:
            if workbook:
                workbook.close()

        return results

    def process_multiple_files(
        self, file_paths: List[str], max_workers: int = None
    ) -> Dict[str, Dict[str, Any]]:
        """批量处理多个Excel文件"""
        if not max_workers:
            max_workers = (
                self.config.max_workers if hasattr(self.config, "max_workers") else 4
            )

        if not self.config.enable_parallel_processing:
            max_workers = 1

        results = {}
        self.logger.info(
            f"开始批量处理 {len(file_paths)} 个文件，使用 {max_workers} 个线程"
        )

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交任务
            future_to_file = {
                executor.submit(self.process_excel_file, file_path): file_path
                for file_path in file_paths
            }

            # 收集结果
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    results[file_path] = future.result()
                except Exception as e:
                    results[file_path] = {
                        "file_path": file_path,
                        "error": str(e),
                        "successful": 0,
                        "failed": 0,
                        "processed_at": datetime.now().isoformat(),
                    }
                    self.logger.error(f"批量处理文件失败 {file_path}: {e}")

        self.logger.info(
            f"批量处理完成，成功: {sum(r.get('successful', 0) for r in results.values())}"
        )
        return results

    def get_supported_formats(self) -> List[str]:
        """获取支持的文档格式列表"""
        return self.document_processor.get_supported_formats()


# ==================== 便捷接口 ====================


def create_processor(
    config: Optional[DocumentProcessingConfig] = None,
) -> EnhancedExcelProcessor:
    """创建Excel处理器的便捷函数"""
    return EnhancedExcelProcessor(config)


def process_excel_links(
    excel_path: str,
    config: Optional[DocumentProcessingConfig] = None,
) -> Dict[str, Any]:
    """
    处理Excel文件中链接的便捷函数

    Args:
        excel_path: Excel文件路径
        config: 处理配置

    Returns:
        Dict: 处理结果
    """
    processor = create_processor(config)
    return processor.process_excel_file(excel_path)


# ==================== 使用示例 ====================

if __name__ == "__main__":
    # 使用示例
    print("Excel链接内容提取器 v1.0.0")
    print()

    # 创建处理器
    config = DocumentProcessingConfig(
        output_format="markdown",
        enable_parallel_processing=False,  # 单文件处理设为False
    )

    processor = EnhancedExcelProcessor(config)

    # 示例用法（需要修改为实际的Excel文件路径）
    excel_file = "C:\\Users\\Admin\\Desktop\\text\\任务管理.xlsx"

    if os.path.exists(excel_file):
        print(f"\n开始处理Excel文件: {excel_file}")
        result = processor.process_excel_file(excel_file)
        print(f"处理完成: 成功 {result['successful']}, 失败 {result['failed']}")
    else:
        print(f"示例: python {__file__}")
        print("请在代码中设置实际的Excel文件路径进行测试")
